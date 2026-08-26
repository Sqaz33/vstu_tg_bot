from __future__ import annotations

from pathlib import Path

import openpyxl
import xlrd

from vstu_schedule_bot.parsing.base import (
    CellBorder,
    CellRegion,
    CellStyle,
    SheetGrid,
    WorkbookGrid,
    WorkbookReader,
)


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\n", " ").split())


def _xls_color(book: xlrd.book.Book, color_index: int, fallback: str) -> str:
    rgb = book.colour_map.get(color_index)
    if not rgb:
        return fallback
    return f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def _xlsx_color(color: object, fallback: str) -> str:
    color_type = getattr(color, "type", None)
    value = getattr(color, "rgb", None) if color_type == "rgb" else None
    if isinstance(value, str):
        return f"#{value[-6:].upper()}"
    return fallback


_BORDER_STYLES = {
    None: 0,
    "hair": 1,
    "thin": 1,
    "dotted": 1,
    "dashDotDot": 1,
    "dashDot": 1,
    "dashed": 1,
    "mediumDashDotDot": 2,
    "slantDashDot": 2,
    "mediumDashDot": 2,
    "mediumDashed": 2,
    "medium": 2,
    "double": 5,
    "thick": 5,
}


def _xlsx_border_style(value: str | None) -> int:
    return _BORDER_STYLES.get(value, 1 if value else 0)


def _xls_cell_style(book: xlrd.book.Book, xf: xlrd.formatting.XF) -> CellStyle:
    border = xf.border
    cell_border = (
        CellBorder(
            top=border.top_line_style,
            bottom=border.bottom_line_style,
            left=border.left_line_style,
            right=border.right_line_style,
        )
        if border is not None
        else CellBorder()
    )
    background = xf.background
    fill_color = (
        _xls_color(book, background.pattern_colour_index, "#FFFFFF")
        if background is not None
        else "#FFFFFF"
    )
    font = book.font_list[xf.font_index]
    return CellStyle(
        border=cell_border,
        font_color=_xls_color(book, font.colour_index, "#000000"),
        fill_color=fill_color,
        bold=bool(font.bold),
    )


class XlsReader:
    def supports(self, suffix: str) -> bool:
        return suffix.lower() == ".xls"

    def read(self, path: Path) -> WorkbookGrid:
        # ``formatting_info`` is needed for BIFF merged-cell metadata. Without it,
        # the top-left values survive but lesson duration and shared group bands do not.
        book = xlrd.open_workbook(str(path), formatting_info=True, on_demand=True)
        sheets: list[SheetGrid] = []
        try:
            for source in book.sheets():
                values = tuple(
                    tuple(_text(source.cell_value(row, col)) for col in range(source.ncols))
                    for row in range(source.nrows)
                )
                merged = {
                    (r0, r1, c0, c1): _text(source.cell_value(r0, c0))
                    for r0, r1, c0, c1 in source.merged_cells
                }
                regions = [CellRegion(*coords, value) for coords, value in merged.items()]
                styles = tuple(
                    tuple(
                        _xls_cell_style(book, xf)
                        for col in range(source.ncols)
                        for xf in (book.xf_list[source.cell_xf_index(row, col)],)
                    )
                    for row in range(source.nrows)
                )
                merged_covered = {
                    (row, col)
                    for r0, r1, c0, c1 in merged
                    for row in range(r0, r1)
                    for col in range(c0, c1)
                }
                for row in range(source.nrows):
                    for col in range(source.ncols):
                        value = values[row][col]
                        if value and (row, col) not in merged_covered:
                            regions.append(CellRegion(row, row + 1, col, col + 1, value))
                sheets.append(
                    SheetGrid(
                        name=source.name,
                        rows=source.nrows,
                        cols=source.ncols,
                        values=values,
                        regions=tuple(regions),
                        styles=styles,
                    )
                )
        finally:
            book.release_resources()
        return WorkbookGrid(filename=path.name, sheets=tuple(sheets))


class XlsxReader:
    def supports(self, suffix: str) -> bool:
        return suffix.lower() in {".xlsx", ".xlsm"}

    def read(self, path: Path) -> WorkbookGrid:
        book = openpyxl.load_workbook(path, read_only=False, data_only=True)
        sheets: list[SheetGrid] = []
        for source in book.worksheets:
            rows, cols = source.max_row, source.max_column
            values = tuple(
                tuple(_text(source.cell(row + 1, col + 1).value) for col in range(cols))
                for row in range(rows)
            )
            merged_coords: dict[tuple[int, int, int, int], str] = {}
            for merged in source.merged_cells.ranges:
                coords = (
                    merged.min_row - 1,
                    merged.max_row,
                    merged.min_col - 1,
                    merged.max_col,
                )
                merged_coords[coords] = values[coords[0]][coords[2]]
            regions = [CellRegion(*coords, value) for coords, value in merged_coords.items()]
            styles = tuple(
                tuple(
                    CellStyle(
                        border=CellBorder(
                            top=_xlsx_border_style(cell.border.top.style),
                            bottom=_xlsx_border_style(cell.border.bottom.style),
                            left=_xlsx_border_style(cell.border.left.style),
                            right=_xlsx_border_style(cell.border.right.style),
                        ),
                        font_color=_xlsx_color(cell.font.color, "#000000"),
                        fill_color=_xlsx_color(cell.fill.fgColor, "#FFFFFF"),
                        bold=bool(cell.font.bold),
                    )
                    for col in range(cols)
                    for cell in (source.cell(row + 1, col + 1),)
                )
                for row in range(rows)
            )
            merged_covered = {
                (row, col)
                for r0, r1, c0, c1 in merged_coords
                for row in range(r0, r1)
                for col in range(c0, c1)
            }
            for row in range(rows):
                for col in range(cols):
                    value = values[row][col]
                    if value and (row, col) not in merged_covered:
                        regions.append(CellRegion(row, row + 1, col, col + 1, value))
            sheets.append(
                SheetGrid(
                    name=source.title,
                    rows=rows,
                    cols=cols,
                    values=values,
                    regions=tuple(regions),
                    styles=styles,
                )
            )
        book.close()
        return WorkbookGrid(filename=path.name, sheets=tuple(sheets))


class WorkbookReaderRegistry:
    def __init__(self, readers: list[WorkbookReader] | None = None) -> None:
        self._readers = readers or [XlsReader(), XlsxReader()]

    def read(self, path: Path) -> WorkbookGrid:
        for reader in self._readers:
            if reader.supports(path.suffix):
                return reader.read(path)
        raise ValueError(f"Unsupported spreadsheet format: {path.suffix}")
